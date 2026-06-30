import os
from decimal import Decimal
from datetime import datetime, date
from pathlib import Path

import django
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "projectdata.settings")
django.setup()

from django.db import transaction
from vardhaman.models import User, Products, Order, Order_data, order_taxes, Esstimate

EXCEL_PATH = Path(__file__).resolve().parent / "vardhamanDB_export.xlsx"

print(start_time := datetime.now())

def parse_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal, bool)):
        return value
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        return text
    return value


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            try:
                return datetime.strptime(text, "%Y-%m-%d").date()
            except ValueError:
                return None
    return None


def parse_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            try:
                return datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None
    return None


def normalize_row(row):
    return [parse_value(v) for v in row]


def load_sheet(sheet_name):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        data_rows.append(dict(zip(headers, normalize_row(row))))
    return data_rows


def import_users():
    rows = load_sheet("users")
    created = 0
    skipped = 0
    for row in rows:
        existing = User.objects.filter(id=row.get("id")).first()
        if existing:
            skipped += 1
            continue
        newUser = User.objects.create(
            id=row.get("id"),
            role=str(row.get("role") or "0"),
            name=row.get("name"),
            shopname=row.get("shopname"),
            email=row.get("email"),
            contact_no=row.get("contact_no"),
            password=row.get("password"),
            register_by=row.get("register_by"),
            is_approved=str(row.get("is_approved") or "0"),
            address=row.get("address"),
            expo_go_token=row.get("expo_go_token"),
            gst_no=row.get("gst_no"),
            bankname=row.get("bankname"),
            holderName=row.get("holderName"),
            branch_name=row.get("branch_name"),
            ifsc_code=row.get("ifsc_code"),
            account_num=row.get("account_num"),
            fassai=row.get("fassai"),
            created_at=parse_datetime(row.get("created_at")) or User._meta.get_field("created_at").default,
            updated_at=parse_datetime(row.get("updated_at")) or User._meta.get_field("updated_at").default,
        )
        print(f"Created user: {newUser.name} (ID: {newUser.id})")
        created += 1
    print(f"Imported {created} users, skipped {skipped} existing")


def import_products():
    rows = load_sheet("Products")
    created = 0
    skipped = 0
    for row in rows:
        existing = Products.objects.filter(id=row.get("id")).first()
        if existing:
            skipped += 1
            continue
        prd = Products.objects.create(
            id=row.get("id"),
            product_name_eng=row.get("product_name_eng"),
            product_name_guj=row.get("product_name_guj"),
            product_name_hin=row.get("product_name_hin"),
            product_image=row.get("product_image"),
            product_qty=row.get("product_qty"),
            product_unit=row.get("product_unit"),
            product_gst_rate=int(row.get("product_gst_rate") or 0),
            product_discount_rate=int(row.get("product_discount_rate") or 0),
            product_price=Decimal(str(row.get("product_price") or 0)),
            product_gst=Decimal(str(row.get("product_gst") or 0)),
            product_tax_price=Decimal(str(row.get("product_tax_price") or 0)),
            product_hsn_code=row.get("product_hsn_code"),
            is_delete=str(row.get("is_delete") or "0"),
            created_at=parse_datetime(row.get("created_at")) or Products._meta.get_field("created_at").default,
            updated_at=parse_datetime(row.get("updated_at")) or Products._meta.get_field("updated_at").default,
        )
        print(f"Created product: {prd.product_name_eng} (ID: {prd.id})")
        created += 1
    print(f"Imported {created} products, skipped {skipped} existing")


def import_orders():
    rows = load_sheet("Orders")
    created = 0
    skipped = 0
    for row in rows:
        existing = Order.objects.filter(id=row.get("id")).first()
        if existing:
            skipped += 1
            continue
        customer = None
        customer_id = row.get("customer_id_id")
        if customer_id:
            customer = User.objects.filter(id=customer_id).first()
        order = Order.objects.create(
            id=row.get("id"),
            order_id=int(row.get("order_id") or 0),
            customer_id=customer,
            name=row.get("name"),
            email=row.get("email"),
            contact_no=row.get("contact_no"),
            gst_no=row.get("gst_no"),
            address=row.get("address"),
            round_off=Decimal(str(row.get("round_off") or 0)),
            round_type=row.get("round_type") or "less",
            total_amount=Decimal(str(row.get("total_amount") or 0)),
            grand_total_amount=Decimal(str(row.get("grand_total_amount") or 0)),
            taxable_amount=Decimal(str(row.get("taxable_amount") or 0)),
            cgst_amount=Decimal(str(row.get("cgst_amount") or 0)),
            sgst_amount=Decimal(str(row.get("sgst_amount") or 0)),
            total_tax_amount=Decimal(str(row.get("total_tax_amount") or 0)),
            status=str(row.get("status") or "0"),
            invoice_date=parse_date(row.get("invoice_date")),
            created_at=parse_datetime(row.get("created_at")),
        )
        print(f"Created order: {order.name} (ID: {order.id})")
        created += 1
    print(f"Imported {created} orders, skipped {skipped} existing")


def import_order_lines():
    rows = load_sheet("Orders_list")
    created = 0
    skipped = 0
    for row in rows:
        existing = Order_data.objects.filter(id=row.get("id")).first()
        if existing:
            skipped += 1
            continue
        order = None
        product = None
        if row.get("order_id_id"):
            order = Order.objects.filter(id=row.get("order_id_id")).first()
        if row.get("product_id_id"):
            product = Products.objects.filter(id=row.get("product_id_id")).first()
        order_line = Order_data.objects.create(
            id=row.get("id"),
            order_id=order,
            product_id=product,
            qty=row.get("qty"),
            cgst_rate=row.get("cgst_rate") or "0",
            sgst_rate=row.get("sgst_rate") or "0",
            cgst_amount=Decimal(str(row.get("cgst_amount") or 0)),
            sgst_amount=Decimal(str(row.get("sgst_amount") or 0)),
            amount=Decimal(str(row.get("amount") or 0)),
            tax_amount=Decimal(str(row.get("tax_amount") or 0)),
            status=str(row.get("status") or "0"),
            created_at=parse_datetime(row.get("created_at")),
        )
        print(f"Created order line: {order_line.id}")
        created += 1
    print(f"Imported {created} order lines, skipped {skipped} existing")


def import_order_taxes():
    rows = load_sheet("vardhaman_order_taxes")
    created = 0
    skipped = 0
    for row in rows:
        existing = order_taxes.objects.filter(id=row.get("id")).first()
        if existing:
            skipped += 1
            continue
        order = None
        if row.get("order_id_id"):
            order = Order.objects.filter(id=row.get("order_id_id")).first()
        order_tax = order_taxes.objects.create(
            id=row.get("id"),
            order_id=order,
            tax_rate=row.get("tax_rate") or "0",
            taxable_amount=Decimal(str(row.get("taxable_amount") or 0)),
            cgst_amount=Decimal(str(row.get("cgst_amount") or 0)),
            sgst_amount=Decimal(str(row.get("sgst_amount") or 0)),
            total_tax_amount=Decimal(str(row.get("total_tax_amount") or 0)),
        )
        print(f"Created order tax: {order_tax.id}")
        created += 1
    print(f"Imported {created} order taxes, skipped {skipped} existing")


def import_estimates():
    rows = load_sheet("vardhaman_esstimate")
    created = 0
    skipped = 0
    for row in rows:
        existing = Esstimate.objects.filter(id=row.get("id")).first()
        if existing:
            skipped += 1
            continue
        esstimate = Esstimate.objects.create(
            id=row.get("id"),
            customer_name=row.get("customer_name"),
            customer_number=row.get("customer_number"),
            invoice_date=parse_date(row.get("invoice_date")),
            items_list=row.get("items_list"),
            grand_total_amount=Decimal(str(row.get("grand_total_amount") or 0)),
        )
        print(f"Created estimate: {esstimate.customer_name} (ID: {esstimate.id})")
        created += 1
    print(f"Imported {created} estimates, skipped {skipped} existing")


def run_restore():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    try:
        with transaction.atomic():
            import_users()
            import_products()
            import_orders()
            import_order_lines()
            import_order_taxes()
            import_estimates()
    except Exception as exc:
        print(f"Restore failed: {exc}")
        raise

    print("Restore completed")
    return {
        "excel_path": str(EXCEL_PATH),
        "status": "completed",
    }


def main():
    run_restore()


if __name__ == "__main__":
    main()
